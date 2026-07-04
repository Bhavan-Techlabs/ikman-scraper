import logging
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from data_processor import COLUMNS, COL, compute_value_score

logger = logging.getLogger(__name__)

OUTPUT_DIR = "output"

# Colours
_HEADER_FILL   = PatternFill("solid", fgColor="2E4057")   # dark navy
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL      = PatternFill("solid", fgColor="F4F8FB")   # very light blue
_OVERVIEW_HEADER_FILL = PatternFill("solid", fgColor="1A3A4A")
_GOLD_FILL     = PatternFill("solid", fgColor="FFF0B3")   # top 3 score rows
_GREEN_FILL    = PatternFill("solid", fgColor="D6F5E3")
_WRAP          = Alignment(wrap_text=True, vertical="top")
_CENTER        = Alignment(horizontal="center", vertical="center")
_TOP           = Alignment(vertical="top")
_THIN          = Side(style="thin", color="CCCCCC")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Excel number formats
_FMT_INT        = "0"
_FMT_FLOAT1     = "0.0"
_FMT_FLOAT2     = "0.00"
_FMT_CURRENCY   = '"Rs."#,##0'

# Column widths  (col name → width chars)
_COL_WIDTHS = {
    "Title":                  40,
    "Location":               14,
    "Bedrooms":               10,
    "Bathrooms":              10,
    "Land Size (Perches)":    18,
    "House Size (SqFt)":      16,
    "Price (LKR)":            18,
    "Address":                30,
    "Description":            50,
    "URL":                    20,
    "Posted":                 12,
    "Date Scraped":           16,
    "Status":                 12,
    "Notes":                  25,
}


def _output_path():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    return os.path.join(OUTPUT_DIR, f"srilanka_house_sales_{date_str}.xlsx")


get_output_path = _output_path


def _get_or_create_workbook(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _format_header_row(ws, col_count):
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font   = _HEADER_FONT
        cell.fill   = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER


def _apply_column_formats(ws, col_names):
    """Set number formats, widths, and alignment for each data column."""
    for col_idx, col_name in enumerate(col_names, 1):
        letter = get_column_letter(col_idx)

        # Width
        ws.column_dimensions[letter].width = _COL_WIDTHS.get(col_name, 14)

        # Number formats & alignment for data rows (row 2 onward)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _BORDER

            if col_name in ("Bedrooms", "Bathrooms"):
                cell.number_format = _FMT_INT
                cell.alignment = _CENTER
            elif col_name == "Land Size (Perches)":
                cell.number_format = _FMT_FLOAT1
                cell.alignment = _CENTER
            elif col_name == "House Size (SqFt)":
                cell.number_format = _FMT_FLOAT2
                cell.alignment = _CENTER
            elif col_name == "Price (LKR)":
                cell.number_format = _FMT_CURRENCY
                cell.alignment = _TOP
            elif col_name in ("Description", "Notes"):
                cell.alignment = _WRAP
            elif col_name == "URL":
                cell.alignment = _TOP
                # Make URL a clickable hyperlink
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = str(cell.value)
                    cell.value = "View Listing"
                    cell.font = Font(color="0563C1", underline="single")
            else:
                cell.alignment = _TOP


def _stripe_rows(ws):
    """Alternate row shading starting at row 2."""
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
                    cell.fill = _ALT_FILL


def _get_or_create_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.append(COLUMNS)
    _format_header_row(ws, len(COLUMNS))
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    return ws


def get_existing_data(path, sheet_name):
    """Read existing rows from the sheet as list of dicts."""
    if not os.path.exists(path):
        return []
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= 1:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def append_rows(sheet_name, new_rows):
    """Append new_rows (list of native-typed values) to the given sheet."""
    if not new_rows:
        logger.info("No new rows to append to Excel for %s", sheet_name)
        return

    path = _output_path()
    wb   = _get_or_create_workbook(path)
    ws   = _get_or_create_sheet(wb, sheet_name)

    for row in new_rows:
        ws.append(row)

    _apply_column_formats(ws, COLUMNS)
    _stripe_rows(ws)
    _rebuild_overview(wb)
    wb.save(path)
    logger.info("Appended %d rows to Excel sheet '%s' → %s", len(new_rows), sheet_name, path)


def clear_sheet(sheet_name):
    """Clear all data rows, keeping the header."""
    path = _output_path()
    if not os.path.exists(path):
        return
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    for row in list(ws.iter_rows(min_row=2)):
        for cell in row:
            cell.value = None
    _rebuild_overview(wb)
    wb.save(path)
    logger.info("Cleared Excel sheet: %s", sheet_name)


# ---------------------------------------------------------------------------
# Overview tab
# ---------------------------------------------------------------------------

_OV_COLUMNS = [
    "Rank",
    "Score",
    "Title",
    "Location",
    "Price (LKR)",
    "Land Size (Perches)",
    "House Size (SqFt)",
    "Bedrooms",
    "Bathrooms",
    "Value Insight",
    "URL",
]

_OV_WIDTHS = {
    "Rank": 6, "Score": 9, "Title": 40, "Location": 14,
    "Price (LKR)": 18, "Land Size (Perches)": 18, "House Size (SqFt)": 16,
    "Bedrooms": 10, "Bathrooms": 10, "Value Insight": 55, "URL": 20,
}

_OV_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)


def _value_insight(score, price, land, house, beds, baths):
    """Human-readable one-line explanation of the value score."""
    parts = []
    if land:
        parts.append(f"{land:.1f}P land")
    if house:
        parts.append(f"{house:.0f} sqft built")
    if beds:
        parts.append(f"{int(beds)}BR")
    if baths:
        parts.append(f"{int(baths)}BA")
    space_desc = " · ".join(parts) if parts else "limited data"
    if price:
        lkr_m = price / 1_000_000
        return f"Rs.{lkr_m:.1f}M — {space_desc} — score {score:.4f}"
    return space_desc


def _rebuild_overview(wb):
    """Rebuild the Overview sheet from all location sheets in the workbook."""
    OVERVIEW_NAME = "Overview"

    # Remove old overview so we rebuild clean
    if OVERVIEW_NAME in wb.sheetnames:
        del wb[OVERVIEW_NAME]

    ws_ov = wb.create_sheet(title=OVERVIEW_NAME, index=0)

    # --- Title banner ---
    ws_ov.merge_cells("A1:K1")
    banner = ws_ov["A1"]
    banner.value = "Sri Lanka House Sales — Value Overview"
    banner.font  = Font(bold=True, color="FFFFFF", size=13)
    banner.fill  = PatternFill("solid", fgColor="1A3A4A")
    banner.alignment = _CENTER
    ws_ov.row_dimensions[1].height = 28

    ws_ov.merge_cells("A2:K2")
    sub = ws_ov["A2"]
    sub.value = (
        "Ranked by value score = (land sqft + built sqft + room bonus) ÷ price × 1M  |  "
        "Higher score = more space per rupee"
    )
    sub.font  = Font(italic=True, color="555555", size=9)
    sub.fill  = PatternFill("solid", fgColor="EEF3F7")
    sub.alignment = _CENTER
    ws_ov.row_dimensions[2].height = 16

    # --- Column headers row 3 ---
    ws_ov.append([])   # blank placeholder, will replace below
    header_row = 3
    for col_idx, col_name in enumerate(_OV_COLUMNS, 1):
        cell = ws_ov.cell(row=header_row, column=col_idx, value=col_name)
        cell.font      = _OV_HEADER_FONT
        cell.fill      = _OVERVIEW_HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
        ws_ov.column_dimensions[get_column_letter(col_idx)].width = _OV_WIDTHS.get(col_name, 14)
    ws_ov.row_dimensions[header_row].height = 22
    ws_ov.freeze_panes = "A4"

    # --- Collect scored rows from every location sheet ---
    scored = []
    for sheet_name in wb.sheetnames:
        if sheet_name == OVERVIEW_NAME:
            continue
        ws_src = wb[sheet_name]
        src_rows = list(ws_src.iter_rows(values_only=True))
        if len(src_rows) <= 1:
            continue
        headers = src_rows[0]
        h = {name: i for i, name in enumerate(headers) if name}

        for data_row in src_rows[1:]:
            def _get(col_name):
                idx = h.get(col_name)
                return data_row[idx] if idx is not None else None

            price  = _get("Price (LKR)")
            land   = _get("Land Size (Perches)")
            house  = _get("House Size (SqFt)")
            beds   = _get("Bedrooms")
            baths  = _get("Bathrooms")
            title  = _get("Title") or ""
            url    = _get("URL") or ""
            loc    = _get("Location") or sheet_name

            # Normalise — values may be stored as native numbers or legacy strings
            def _num(v):
                if v is None or v == "":
                    return None
                try:
                    return float(str(v).replace(",", ""))
                except (ValueError, TypeError):
                    return None

            price_n = _num(price)
            land_n  = _num(land)
            house_n = _num(house)
            beds_n  = _num(beds)
            baths_n = _num(baths)

            score = compute_value_score(price_n, land_n, house_n, beds_n, baths_n)
            if score is None:
                continue

            scored.append({
                "score": score,
                "title": title,
                "location": loc,
                "price": price_n,
                "land": land_n,
                "house": house_n,
                "beds": beds_n,
                "baths": baths_n,
                "url": url,
            })

    scored.sort(key=lambda r: r["score"], reverse=True)

    # --- Write rows ---
    for rank, r in enumerate(scored, 1):
        insight = _value_insight(r["score"], r["price"], r["land"], r["house"], r["beds"], r["baths"])
        row_data = [
            rank,
            r["score"],
            r["title"],
            r["location"],
            r["price"],
            r["land"],
            r["house"],
            int(r["beds"]) if r["beds"] is not None else None,
            int(r["baths"]) if r["baths"] is not None else None,
            insight,
            r["url"],
        ]
        ws_ov.append(row_data)
        data_row_idx = ws_ov.max_row

        # Row fill: gold for top 3, alternating light blue after
        if rank <= 3:
            fill = _GOLD_FILL
        elif rank % 2 == 0:
            fill = _ALT_FILL
        else:
            fill = PatternFill()  # default white

        for col_idx, col_name in enumerate(_OV_COLUMNS, 1):
            cell = ws_ov.cell(row=data_row_idx, column=col_idx)
            cell.border = _BORDER

            if fill.fill_type:
                cell.fill = fill

            if col_name == "Rank":
                cell.alignment = _CENTER
                cell.font = Font(bold=(rank <= 3))
            elif col_name == "Score":
                cell.number_format = "0.0000"
                cell.alignment = _CENTER
            elif col_name == "Price (LKR)":
                cell.number_format = _FMT_CURRENCY
                cell.alignment = _TOP
            elif col_name in ("Land Size (Perches)",):
                cell.number_format = _FMT_FLOAT1
                cell.alignment = _CENTER
            elif col_name in ("House Size (SqFt)",):
                cell.number_format = _FMT_FLOAT2
                cell.alignment = _CENTER
            elif col_name in ("Bedrooms", "Bathrooms"):
                cell.number_format = _FMT_INT
                cell.alignment = _CENTER
            elif col_name == "Value Insight":
                cell.alignment = _WRAP
            elif col_name == "URL":
                if r["url"] and r["url"].startswith("http"):
                    cell.hyperlink = r["url"]
                    cell.value = "View Listing"
                    cell.font = Font(color="0563C1", underline="single")
                cell.alignment = _TOP
            else:
                cell.alignment = _TOP

    # Row heights for data rows
    for row_idx in range(4, ws_ov.max_row + 1):
        ws_ov.row_dimensions[row_idx].height = 30

    # --- Summary stats block below the table ---
    gap_row = ws_ov.max_row + 2
    ws_ov.cell(row=gap_row, column=1, value="Summary").font = Font(bold=True)

    if scored:
        avg_score = sum(r["score"] for r in scored) / len(scored)
        ws_ov.cell(row=gap_row + 1, column=1, value=f"Total listings scored: {len(scored)}")
        ws_ov.cell(row=gap_row + 2, column=1, value=f"Average score: {avg_score:.4f}")
        ws_ov.cell(row=gap_row + 3, column=1, value=f"Best value: {scored[0]['title'][:60]}")

    logger.info("Overview tab rebuilt with %d scored listings", len(scored))
